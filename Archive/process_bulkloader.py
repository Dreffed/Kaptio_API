import json

from kaptiorestpython.client import KaptioClient, load_kaptioconfig
from kaptiorestpython.ograph import KaptioOGraph
from utils import get_pickle_data, save_pickle_data, save_json, scanfiles

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

import string
import pickle
import os
from time import time
from datetime import datetime, timedelta
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def load_WB(path):    
    # process each sheet...
    # assume top row is the titles
    # subsequent rows are data...
    wb = load_workbook(path)
    return wb


homepath = os.path.expanduser("~")
datapaths = ["OneDrive - Great Canadian Railtour Co", "Jupyter_NB"]
savepath = os.path.join(homepath, *datapaths)
logger.info(savepath)

kaptio_config_file = os.path.join(savepath, "config", "kaptio_settings.json")
kaptio_config = load_kaptioconfig(kaptio_config_file)

configxl_path = os.path.join(savepath, 'config', 'all_sell.xl.config.json')

debug = True
baseurl = kaptio_config['ograph']['baseurl']
sfurl = kaptio_config['sf']['url']
username = kaptio_config['sf']['username']
password = kaptio_config['sf']['passwd']
security_token = kaptio_config['sf']['token']
sandbox = True
clientid = kaptio_config['ograph']['clientid']
clientsecret = kaptio_config['ograph']['clientsecret']

kt = KaptioOGraph(baseurl, sfurl, username, password, security_token, sandbox, clientid, clientsecret)

pickle_file = "kaptio_allsell.pickle"
data = get_pickle_data(pickle_file)

logger.info("Available items:")
for key, value in data.items():
    logger.info("\t{}".format(key))

packageid = 'a754F0000000A30QAE'
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
logger.info("Timestamp: {}".format(timestamp))

tax_profile = 'Domestic' #'Foreign' #'Zero Rated'
tax_profiles = data.get('tax_profiles', {})
for key, value in tax_profiles.items():
    starred = " "
    if key == tax_profile:
        starred = "*"
    logger.info("\t{}{} => {}".format(starred, key, value))

occupancy = data['occupancy']
search_values = data['search_values']
season_start = data['season']['start']
season_end = data['season']['end']
kt_packages = data['packages']
kt_pricelist = data['pricelist']

# load the data...
packagerows = []
for p in kt_packages:
    row = {}
    if p.get('active') != True:
        if p.get('id') not in config.get('packages',[]):  
            continue
        logger.info('\t==> including {}'.format(p.get('name')))

    row['packageid'] = p['id']
    row['packagename'] = p['name']
    row['packagelength'] = p['length']
    row['packagedeptype'] = p['departure_type_id']

    row['packagestart'] = 'missing'
    if 'start_location' in p:
        if p['start_location'] is not None:
            row['packagestart'] = p['start_location']['name']

    packagerows.append(row)
logger.info("Found {} packages".format(len(packagerows)))

error_list = []
price_data = []
for row in packagerows:
    packageid = row['packageid']
    if packageid in kt_pricelist:
        p_data = kt_pricelist[packageid]
        service_levels = {}
        if 'service_levels' in p_data:
            for item in p_data['service_levels']:
                sid = item['id']
                sname = item['name']
                sactive = item['active']
                if not sid in service_levels:
                    service_levels[sid] = {'name': sname, 'active': sactive}

        if 'pricelist' in p_data:
            for d_key, d_value in p_data['pricelist'].items():
                if d_key == 'errors':
                    continue

                # process the date
                date_str = d_key
                p_date = datetime.strptime(date_str, "%Y-%m-%d")
                dow = p_date.strftime('%a')
                try:
                    e_date = p_date + timedelta(days=row['packagelength'])
                    e_date_str = e_date.strftime('%Y-%m-%d')
                except:
                    e_date = p_date

                b = {}
                b['packageid'] = row['packageid']
                b['packagename'] = row['packagename']
                b['packagedeptype'] = row['packagedeptype']
                b['depdate'] = p_date
                b['arrdate'] = e_date
                b['mon'] = ('Y' if dow =='Mon' else '')
                b['tue'] = ('Y' if dow =='Tue' else '')
                b['wed'] = ('Y' if dow =='Wed' else '')
                b['thu'] = ('Y' if dow =='Thu' else '')
                b['fri'] = ('Y' if dow =='Fri' else '')
                b['sat'] = ('Y' if dow =='Sat' else '')
                b['sun'] = ('Y' if dow =='Sun' else '')

                for t_key, t_value in d_value.items():
                    if t_key != tax_profile:
                        continue
                    o_data  = {}
                    for o_key, o_value in t_value.items():
                        # now we are in the array of service types...
                        
                        o_data[o_key] = []

                        for item in o_value:
                            try:
                                if len(item['errors']) == 0:

                                    if 'total_price' in item:
                                        t = item['total_price']
                                        r = {}
                                        r['service_level_id'] = item['service_level_id']
                                        r['service_level'] = service_levels[item['service_level_id']]
                                        r['tax_profile'] = t_key
                                        r["net"] = t["net"]
                                        r['sales'] = t["sales"]
                                        r['net_discount'] = t["net_discount"]
                                        r['sales_discount'] = t["sales_discount"]
                                        r['tax'] = t["tax"]
                                        r['currency'] = t["currency"]
                                        r['supplier_price'] = t["supplier_price"]
                                        r = {**r}
                                        o_data[o_key].append(r)
                                else:
                                    for err in item['errors']:
                                        r_err = {}
                                        r_err['packageid'] = row['packageid']
                                        r_err['packagename'] = row['packagename']
                                        r_err['packagedeptype'] = row['packagedeptype']
                                        r_err['depdate'] = d_key
                                        r_err['tax_profile'] = t_key
                                        r_err['service_level_id'] = item['service_level_id']
                                        r_err['service_level'] = service_levels[item['service_level_id']]
                                        r_err['code'] = err['error']['code']
                                        r_err['message'] = err['error']['message']
                                        r_err['details'] = err['error']['details']
                                        error_list.append(r_err)
                            except Exception as ex:
                                logger.info('Error: {}'.format(ex))
                                logger.info(json.dumps(row, indent=4))
                                logger.info(json.dumps(o_value, indent=4))
                    # pivot the occupancy data...
                    # we should have a dict with key of occupancy key and a row for each service level...
                    for s_key, s_value in service_levels.items():
                        s = {}
                        s['service_level_id'] = s_key
                        s['service_level'] = s_value['name']
                        for o_key, o_value in o_data.items():
                            for item in o_value:
                                if item['service_level_id'] == s_key:
                                    # we want this data...
                                    s_prefix = 'xx'
                                    if o_key == 'single':
                                        s_prefix = 'sa'
                                    elif o_key == 'double':
                                        s_prefix = 'da'
                                    elif o_key == 'triple':
                                        s_prefix = 'ta'
                                    elif o_key == 'quad':
                                        s_prefix = 'qa'

                                    s['{}sales'.format(s_prefix)] = item['sales']
                                    s['{}net'.format(s_prefix)] = item['net']
                                    s['{}tax'.format(s_prefix)] = item['tax']
                                    s['{}currency'.format(s_prefix)] = item['currency']
                                    s['tax_profile'] = item['tax_profile']
                        r = {**b, **s}
                        price_data.append(r)

logger.info("Rows: {} => Errors: {}".format(len(price_data), len(error_list)))


if not os.path.exists(configxl_path):
    excel_config = {
        "fieldmap": [
            {
                "border": "default",
                "column": "H",
                "field": "departureCity",
                "fill": "default",
                "font": "default",
                "name": "city"
            }
        ],
        "fills": {
            "bomap": {
                "fill_type": "solid",
                "start_color": "F2DCDB"
            },
            "dates": {
                "fill_type": "solid",
                "start_color": "F2DCDB"
            },
            "default": {
                "fill_type": "solid",
                "start_color": "FFFFCC"
            },
            "dentry": {
                "fill_type": "solid",
                "start_color": "BFBFBF"
            },
            "kids": {
                "fill_type": "solid",
                "start_color": "FFCC99"
            },
            "lentry": {
                "fill_type": "solid",
                "start_color": "E6E6E6"
            },
            "map": {
                "fill_type": "solid",
                "start_color": "E6E6F1"
            }
        },
        "fonts": {
            "default": {
                "name": "Ariel",
                "size": 18
            },
            "formula": {
                "italic": True,
                "name": "Ariel",
                "size": 18
            }
        },
        "sheetname": "2018",
        "sides": {
            "default": {
                "color": "B2B2B2",
                "style": "thin"
            }
        },
        "start_row": 9
    }
    with open(configxl_path, 'w') as outfile:
        json.dump(excel_config, outfile, indent=4)
    
else:
    # load the settings...
    excel_config = json.load(open(configxl_path))

# now to load the data into the 
year = 2020
version = "1.2"

excel_feed_path = os.path.join(savepath, 'templates', r'Rocky Bulk Cost Loader template.xlsx')
bulk_file_name = 'Rocky Bulk Loader.{}.{}.{}.xlsx'.format(tax_profile.replace(' ',''),year, version)

row_idx = excel_config['start_row']-1

logger.info('=== Styles ===')
style_fills = {}
if 'fills' in excel_config:
    logger.info('\t=== Files ===')
    for key in excel_config['fills']:
        logger.info('\t\t{} -> {}'.format(key, excel_config['fills'][key]))
        style_fills[key] = PatternFill(**excel_config['fills'][key])
        logger.debug(style_fills[key])

style_fonts = {}
if 'fonts' in excel_config:
    logger.info('\t=== Fonts ===')
    for key in excel_config['fonts']:
        logger.info('\t\t{} -> {}'.format(key, excel_config['fonts'][key]))
        style_fonts[key] = Font(**excel_config['fonts'][key])
        logger.debug(style_fonts[key])

style_sides = {}
if 'sides' in excel_config:
    logger.info('\t=== Sides ===')
    for key in excel_config['sides']:
        logger.info('\t\t{} -> {}'.format(key, excel_config['sides'][key]))
        style_sides[key] = Side(**excel_config['sides'][key])
        logger.debug(style_sides[key])

wb = load_WB(excel_feed_path)
ws = wb[excel_config['sheetname']]
ws.title = "{}".format(year)

for row in price_data:
    row_idx += 1 
    for field in excel_config['fieldmap']:
        coord = '{}{}'.format(field['column'], row_idx)
    
        # format the field..
        # format the field..
        key_name = 'fill'
        if key_name in field:
            ws[coord].fill = style_fills[field[key_name]]
            
        key_name = 'font'
        if key_name in field:
            ws[coord].font = style_fonts[field[key_name]]

        key_name = 'border'
        if key_name in field:
            cSide = style_sides[field[key_name]]
            ws[coord].border = Border(left=cSide , right=cSide, top=cSide, bottom=cSide)
            
        # condition is a check for NONE in the specified field is none skip that record...
        clear_cell = False
        if 'condition' in field:
            condition_field = field['condition']
            
            if condition_field in row:
                if row[condition_field] is None or row[condition_field] == '':
                    clear_cell = True
            else:
                clear_cell = True

        if clear_cell:
            ws[coord].value = None
            continue
            
        if 'field' in field:
            key = field['field']
            if key in row:
                if 'format' in field and 'type' in field:
                    field_type = field['type']
                    if field_type == 'date':
                        ws[coord].value = row[key].strftime(field['format'])
                    else:
                        ws[coord].value = row[key]
                else:    
                    ws[coord].value = row[key]
                    
        elif 'formula' in field:
            formula = field['formula']
            
            formula = formula.replace('{r}', str(row_idx))
            ws[coord].value = formula
        
        elif 'equation' in field:
            pass
            
        elif 'value' in field:
            ws[coord].value = field['value']
                       
excel_save_path = os.path.join(savepath, 'data', bulk_file_name)
wb.save(excel_save_path)

logger.info('Document saved... {}'.format(excel_save_path))