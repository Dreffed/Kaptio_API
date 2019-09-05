import string
import json
import os

from time import time
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from Equation import Expression
import logging

logger = logging.getLogger(__name__)

def num(s):
    if isinstance(s, float): 
        return s
    if isinstance(s, int):
        return s

    try:
        if '.' in s:
            return float(s)
        return int(s)
    except ValueError:
        return float(s)
    except:
        return 0

def load_WB(path):    
    # process each sheet...
    # assume top row is the titles
    # subsequent rows are data...
    wb = load_workbook(path)
    return wb

def load_bulkloaderconfig(config_path):
    if not os.path.exists(config_path):
        config = {
            "fieldmap": [
                {
                    "border": "default",
                    "column": "H",
                    "field": "departureCity",
                    "fill": "default",
                    "font": "default",
                    "name": "city"
                }
            ],
            "fills": {
                "bomap": {
                    "fill_type": "solid",
                    "start_color": "F2DCDB"
                },
                "dates": {
                    "fill_type": "solid",
                    "start_color": "F2DCDB"
                },
                "default": {
                    "fill_type": "solid",
                    "start_color": "FFFFCC"
                },
                "dentry": {
                    "fill_type": "solid",
                    "start_color": "BFBFBF"
                },
                "kids": {
                    "fill_type": "solid",
                    "start_color": "FFCC99"
                },
                "lentry": {
                    "fill_type": "solid",
                    "start_color": "E6E6E6"
                },
                "map": {
                    "fill_type": "solid",
                    "start_color": "E6E6F1"
                }
            },
            "fonts": {
                "default": {
                    "name": "Ariel",
                    "size": 18
                },
                "formula": {
                    "italic": True,
                    "name": "Ariel",
                    "size": 18
                }
            },
            "sheetname": "2018",
            "sides": {
                "default": {
                    "color": "B2B2B2",
                    "style": "thin"
                }
            },
            "start_row": 9
        }
        with open(config_path, 'w') as outfile:
            json.dump(config, outfile, indent=4)
        
    else:
        # load the settings...
        config = json.load(open(config_path))

    logger.info("Loaded bulkloader config data...")
    return config

def get_field(field, row):
    key = field['field']
    if key in row:
        if 'format' in field and 'type' in field:
            field_type = field['type']
            if field_type == 'date':
                return row[key].strftime(field['format'])
            else:
                return row[key]
        else:    
            return row[key]

def generate_bulkloader(price_data, data, savepath, template, yearnumber, versionnumber, tax_profile, config, currency="CAD"):
    if template is None:
        template = config.get('template')

    excel_feed_path = os.path.join(savepath, 'templates', template)
    bulk_file_name = 'RM Bulk Loader.{}.{}.{}.{}.xlsx'.format(tax_profile.replace(" ", ""), yearnumber, versionnumber, currency)

    row_idx = config['start_row']-1

    logger.info('=== Styles ===')
    style_fills = {}
    if 'fills' in config:
        logger.info('\t=== Fills ===')
        for key in config['fills']:
            logger.info('\t\t{} -> {}'.format(key, config['fills'][key]))
            style_fills[key] = PatternFill(**config['fills'][key])
            logger.debug(style_fills[key])

    style_fonts = {}
    if 'fonts' in config:
        logger.info('\t=== Fonts ===')
        for key in config['fonts']:
            logger.info('\t\t{} -> {}'.format(key, config['fonts'][key]))
            style_fonts[key] = Font(**config['fonts'][key])
            logger.debug(style_fonts[key])

    style_sides = {}
    if 'sides' in config:
        logger.info('\t=== Borders ===')
        for key in config['sides']:
            logger.info('\t\t{} -> {}'.format(key, config['sides'][key]))
            style_sides[key] = Side(**config['sides'][key])
            logger.debug(style_sides[key])

    wb = load_WB(excel_feed_path)
    ws = wb[config['sheetname']]
    for s in config.get('sheets'):
        ws_new = wb.copy_worksheet(ws)
        ws_new.title = "{}".format(s.get('name'))
    wb.remove(ws)

    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
        wb.remove(ws)
    
    for row in price_data:
        # add in the marketingnames if needed...
        p = data.get('marketingnames', {}).get(row.get('packageid'))
        if not p.get('packagecode'):
            continue

        row_idx += 1 
        #if row_idx > 30:
        #    break

        # augment the data...
        row['packagetitle'] = p.get('packagetitle')
        row['packagebrand'] = p.get('packagebrand')
        row['packagecode'] = p.get('packagecode')
        row['packagecat'] = p.get('packagecat')

        for s in config.get('sheets',{}):
            ws = wb[s.get('name')]
            logger.debug("using sheet: {}".format(ws.title))

            for field in s.get('fieldmap',[]):
                logger.debug("\tfields: {}".format(field))
                coord = '{}{}'.format(field['column'], row_idx)
                # format the field..
                key_name = 'fill'
                if key_name in field:
                    ws[coord].fill = style_fills[field[key_name]]
                    
                key_name = 'font'
                if key_name in field:
                    ws[coord].font = style_fonts[field[key_name]]

                key_name = 'border'
                if key_name in field:
                    cSide = style_sides[field[key_name]]
                    ws[coord].border = Border(left=cSide , right=cSide, top=cSide, bottom=cSide)
                    
                key_name = 'number_format'
                if key_name in field:
                    ws[coord].number_format = field.get('number_format')
                    
                # condition is a check for NONE in the specified field is none skip that record...
                clear_cell = False
                if 'condition' in field:
                    condition_field = field['condition']
                    
                    if condition_field in row:
                        if row[condition_field] is None or row[condition_field] == '':
                            clear_cell = True
                    else:
                        clear_cell = True

                if clear_cell:
                    ws[coord].value = None
                    continue
                    
                if 'field' in field:
                    ws[coord].value = get_field(field, row)

                elif 'equation' in field:
                    # split on spaces...
                    fn = Expression(field.get('equation'))
                    t = {}
                    for f in fn:
                        logger.debug("\t\t{}: [{}|{}]".format(f, row.get(f), s.get('parameters',{}).get(f)))
                        t[f] = num(row.get(f,s.get('parameters',{}).get(f)))
                    logger.debug("params: {}".format(t))
                    ws[coord].value = fn(**t)
                            
                elif 'formula' in field:
                    formula = field['formula']
                    formula = formula.replace('{r}', str(row_idx))
                    ws[coord].value = formula
                                
                elif 'value' in field:
                    ws[coord].value = field['value']

    excel_save_path = os.path.join(savepath, 'data', bulk_file_name)
    wb.save(excel_save_path)

    logger.info('Document saved... {}'.format(excel_save_path))